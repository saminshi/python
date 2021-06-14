'''
@Descripttion: test
@version: 
@Author: samin
Date: 2021-02-19 16:19:07
LastEditors: Please set LastEditors
LastEditTime: 2021-06-14 09:38:03
'''

import openpyxl

workbook = openpyxl.Workbook()
# 创建sheet，0指定位置，默认插在指定位置的后面
workbook.create_sheet('test_case', 0)
sheet=workbook.active
#特定单元写入值
sheet['A1']='hi,wwu'
# sheet.cell(row = 4,column = 2,value='test') ##特定单元格写入值
# sheet.append([1,2,3])    # 可以用列表来写入
# 设定sheet的名字，
sheet.title = 'newTitle'





# 操作单元格
print(workbook['newTitle']['A1'].value) #打印Sheet1的A1单元格的值
print(workbook.sheetnames) #列表打印所有的sheet的name
# sheet.rows 同 sheet.iter_rows() #打印所有行（生成器）
# sheet.columns 同sheet.iter_cols()#打印所有列（生成器）
sheet['A'] #A那一列（元组）
sheet['A:C'] #A到C列
sheet[1]#第1行（元组）
sheet[1:3] #第1到3行（元组）
# sheet.cell(row = 1,column = 1).value 同 sheet[1][1].value #ws[i][j]这种方法j(列号)可以为负索引，且i,j都可以为切片，即[2:3]这种格式


workbook.save('new.xlsx')


# 公式
# ws['A4']  =  "=SUM(A1:A3)"

# 合并单元格
# ws.merge_cells(start_row = 2,start_column = 1,end_row = 2,end_column = 2)

# 插入图片
# from openpyxl.drawing.image import Image
# img = Image(r'.\test.png')
# img_size=(50.50)
# img.width,img.height = img_size #设定图片大小
# ws.column_dimensions[A].width = 100
# ws.row_dimensions[1].height = 100   #修改第A列的宽度和第1行的高度
# ws.add_image(img,'A1')


# 设置样式
# from openpyxl.styles import colors from openpyxl.styles import Font
# style = Font(name='宋体',color=colors.RED,italic=True,size=14) #colors='#AABBCC'也可以，Font的其它属性可以dir(Font)看到
# ws['A1'].font = font #设定指定单元格的样式
# col = ws.column_dimensions['A']
# col.font = style #设定列的样式
# row = ws.row_dimensions[1]
# row.font = style #设定行的样式

# 颜色背景
# from openpyxl.sytels import PatternFill
# sytle = PatternFill('solid',fgColor='#AABBCC',bgColor='#DDEEFF')
# ws['A1'].fill = style #指定单元格颜色
# col = ws.column_dimensions['A']
# col.fill = style #设定列的颜色
# row = ws.row_dimensions[1]
# row.fill = style #设定行的颜色

# 画图
# from openpyxl.chart import Series,LineChart,Reference
# chart = LineChart() #图标对象
# chart.title = '图标标题'
# chart.style = 8  #线的sytle
# data = Reference(ws,min_col=5,min_row=4,max_col=10,max_row=4) #数据
# serieobj = Series(data,title='测试') #series对象
# chart.appen(seriesobj) #series添加到chart中
# ws.add_chart(chart,'A6') #chart添加到sheet中